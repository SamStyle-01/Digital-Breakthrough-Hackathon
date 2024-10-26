import asyncio
import logging
import sys
import speech_recognition as sr
from io import BytesIO
from pydub import AudioSegment

from aiogram import Bot, Dispatcher, types, F, html
from aiogram.client.default import DefaultBotProperties
from aiogram.enums import ParseMode
from aiogram.filters import CommandStart, Command
from aiogram.types import Message
from docx import Document
from openpyxl import load_workbook

TOKEN = # Вставить токен
dp = Dispatcher()
mime_type = None
times = 0


def read_docx(file_content):
    document = Document(file_content)
    text = []
    for paragraph in document.paragraphs:
        text.append(paragraph.text)

    return "\n".join(text)


def read_excel(file_content):
    workbook = load_workbook(filename=file_content)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    return data


@dp.message(F.document)
async def handle_document(message: Message, bot: Bot):
    global mime_type
    document = message.document
    file_name = document.file_name
    mime_type = document.mime_type

    if mime_type in [
        "application/pdf",
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    ]:
        file_data = await bot.download(document.file_id)

        await message.reply(f"Файл '{file_name}' получен и сохранен в переменную.")

        if mime_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
            await message.reply(f"{read_docx(file_data)[:300]}")
        elif mime_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
            ex_file = read_excel(file_data)

            rex_file = ""
            for i in range(len(ex_file)):
                for j in range(len(ex_file[i])):
                    try:
                        rex_file = rex_file + ex_file[i][j] + ", "
                    except TypeError:
                        pass
                rex_file = rex_file + "\n"
            await message.reply(f"{rex_file[:300]}")
    else:
        await message.reply("Пожалуйста, отправьте файл в формате PDF, DOCX или XLSX.")


def save_audio_file(audio_bytes, audio_format="ogg"):
    global times
    times += 1
    output_path = fr"C:\Users\Sam\Desktop\output{times}.wav"
    audio = AudioSegment.from_file(BytesIO(audio_bytes), format=audio_format)
    audio.export(output_path, format="wav")
    print(f"Файл успешно сохранён по пути: {output_path}")


def recognize_speech_from_audio(audio_bytes, audio_format="ogg"):
    audio = AudioSegment.from_file(BytesIO(audio_bytes), format=audio_format)

    wav_io = BytesIO()
    audio.export(wav_io, format="wav")
    wav_io.seek(0)
    save_audio_file(audio_bytes)


    recognizer = sr.Recognizer()
    text = ""
    with sr.AudioFile(wav_io) as source:
        audio_data = recognizer.record(source)
        try:
            text = recognizer.recognize_google(audio_data, language="ru-RU")
            text = f"Распознанный текст: {text}"
        except sr.UnknownValueError:
            text = "Не удалось распознать голос."
        except sr.RequestError:
            text = "Ошибка сервиса распознавания речи."

    return text


@dp.message(F.voice)
async def handle_voice(message: Message, bot: Bot):
    voice = await bot.download(message.voice.file_id)
    audio_bytes = voice.read()
    recognized_text = recognize_speech_from_audio(audio_bytes)
    await message.reply(recognized_text)


@dp.message()
async def echo_handler(message: Message) -> None:
    try:
        await message.send_copy(chat_id=message.chat.id)
    except TypeError:
        await message.answer("Nice try!")


@dp.message(CommandStart())
async def command_start_handler(message: Message):
    await message.answer(
        f"Привет, {message.from_user.full_name}! Отправьте PDF, DOCX или XLSX документ, чтобы я его обработал.")


async def main():
    bot = Bot(token=TOKEN, default=DefaultBotProperties(parse_mode=ParseMode.HTML))
    await dp.start_polling(bot)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, stream=sys.stdout)
    asyncio.run(main())
